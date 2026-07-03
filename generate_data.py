"""Generate data.js for the event-checklist app from the Excel workbook."""
import openpyxl, json, re, datetime

SRC = r"D:\Downloads\Event Checklist BETA (1).xlsx"
OUT = r"C:\Users\curse\projects\event-checklist\data.js"

CHANNELS = [
    ("some",   "SoMe"),
    ("news",   "News"),
    ("events", "Events"),
    ("web",    "Own Webpage"),
    ("poster", "Printed Poster"),
    ("mail",   "E-Mail"),
    ("uni",    "Uni Event Website"),
    ("cover",  "Coverage"),
    ("recap",  "Recap SoMe"),
]
COLS = "CDEFGHIJK"  # channel columns in sheet order

RANGE_RE = re.compile(r"^(\d{1,2})\.\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})$")
SINGLE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$")

def parse_date(v):
    """Return (start_iso, end_iso) from a cell value."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        s = v.strftime("%Y-%m-%d")
        return s, s
    if isinstance(v, str):
        v = v.strip()
        m = RANGE_RE.match(v)
        if m:
            d1, d2, mo, y = (int(x) for x in m.groups())
            return (f"{y:04d}-{mo:02d}-{d1:02d}", f"{y:04d}-{mo:02d}-{d2:02d}")
        m = SINGLE_RE.match(v)
        if m:
            d, mo, y = (int(x) for x in m.groups())
            s = f"{y:04d}-{mo:02d}-{d:02d}"
            return s, s
    return None, None

def event_type(title):
    t = title.lower()
    if "jcmml" in t: return "JCMML Lecture"
    if "workshop" in t: return "Workshop"
    if "conference" in t: return "Conference"
    if "film" in t: return "Film Screening"
    if "book launch" in t or "book talk" in t: return "Book Event"
    if "panel" in t: return "Panel"
    if "exhibition" in t: return "Exhibition"
    if "lecture" in t: return "Lecture"
    if "reading" in t: return "Reading"
    return "Special"

FIXES = {
    "Conference Hermeneutics of Restitution (Weller": "Conference Hermeneutics of Restitution (Weller)",
}

wb = openpyxl.load_workbook(SRC, data_only=True)
data = {}
for ws in wb.worksheets:
    year = ws.title
    events = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=13):
        cells = {c.coordinate[0]: c.value for c in row}
        title = cells.get("B")
        if not title or not str(title).strip():
            continue
        title = FIXES.get(str(title).strip(), str(title).strip())
        start, end = parse_date(cells.get("A"))
        channels = {}
        for key_col, (key, _label) in zip(COLS, CHANNELS):
            v = cells.get(key_col)
            if v is None or (isinstance(v, str) and not v.strip()):
                channels[key] = None          # pending / unassigned
            elif isinstance(v, str) and v.strip() == "-":
                channels[key] = "-"           # not applicable
            else:
                # keep only abbreviated initials (all-caps, e.g. NW, BND); drop
                # full-name entries like "Nina", "Dani", "Aze"
                keep = [t.strip() for t in str(v).split(",")
                        if t.strip().isalpha() and t.strip().isupper()]
                channels[key] = ", ".join(keep) if keep else None
        done_raw = cells.get("L")
        done = isinstance(done_raw, str) and done_raw.strip().lower() == "yes"
        events.append({
            "id": f"y{year}r{row[0].row}",
            "start": start,
            "end": end,
            "title": title,
            "type": event_type(title),
            "channels": channels,
            "done": done,
        })
    data[year] = events

people = set()
for evs in data.values():
    for e in evs:
        for v in e["channels"].values():
            if v and v != "-":
                for p in v.split(","):
                    people.add(p.strip())

payload = {
    "channels": [{"key": k, "label": l} for k, l in CHANNELS],
    "people": sorted(people),
    "years": data,
}
with open(OUT, "w", encoding="utf-8") as f:
    f.write("// Seed data imported from 'Event Checklist BETA.xlsx'\n")
    f.write("window.SEED_DATA = ")
    json.dump(payload, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print("people:", sorted(people))
for y, evs in data.items():
    print(y, len(evs), "events,", sum(e["done"] for e in evs), "done,",
          sum(1 for e in evs if e["start"] is None), "unparsed dates")
